
import streamlit as st
from pdf2image import convert_from_path
import fitz as pymupdf
import fitz
import numpy as np
import pandas as pd
import pytesseract
import warnings
warnings.filterwarnings("ignore", category=FutureWarning) 
import unicodedata
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def process_num(x):
    for a in range(len(x)): 
        if x[a] == '':
            pass
        else:
            if '(' in x[a]:
                if ')' in x[a]:
                    x[a] = x[a].strip('(').strip(')')
                    x[a] = '-'+x[a]

            x[a] = x[a].replace(',','')
            x[a] = x[a].replace('.','')
            x[a] = x[a].replace(' ','')

            if x[a][0] == '0':
                x[a] = 0
            
            if x[a] in ['-','—',"=",'_','--']:
                x[a] = 0
            try:
                x[a] = int(x[a])
            except:
                x[a] = 'CHECK'
        
    return x

def identify_separators(col):
    '''Checks a column of data to identify the symbols being used for thousand separation and/or decimal places
     This is done by checking the number of figures between these symbols or the number of figures after the symbol before the end'''
    comma_spaces = [len(x.strip(')').strip('(').split(',')[-1]) for x in col if ',' in x]
    dp_spaces = [len(x.strip(')').strip('(').split('.')[-1]) for x in col if '.' in x]
    average_comma_spaces = np.median(comma_spaces)
    average_dp_spaces = np.median(dp_spaces)
    KSep = 'None'
    DSep = 'None'

    if len(col) != 0:
        if average_comma_spaces != np.nan and len(comma_spaces)/len(col) > 0.1:
            if average_comma_spaces ==3:
                KSep = ','
            elif average_comma_spaces <3:
                DSep = ','

        if average_dp_spaces != np.nan and len(dp_spaces)/len(col) > 0.1:
            if average_dp_spaces ==3:
                KSep = '.'
            elif average_dp_spaces <3:
                DSep = '.'

        if sum([(',' in item and '.' in item) for item in col])/len(col) > 0:
            if DSep == ',':
                KSep = '.'
            if DSep == '.':
                KSep = ','

            #combine both

    if DSep != 'None' and DSep != KSep:
        DP = 'Yes'
    else:
        DP = 'No'

    return DP, KSep

def process_num_dp(x,KSep):
    ''' Slight variation of the normal number processing function to be used when there is both a thousands and decimal point separator'''
    for a in range(len(x)): 
        if x[a] == '':
            pass
        else:
            if '(' in x[a]:
                if ')' in x[a]:
                    x[a] = x[a].strip('(').strip(')')
                    x[a] = '-'+x[a]

            #x[a] = x[a].replace(',','')
            x[a] = x[a].replace(KSep,'')
            x[a] = x[a].replace(',','.') #once the thousands are gone, we can make them floats
            x[a] = x[a].replace(' ','')

            if x[a][0] == '0':
                x[a] = 0
            
            if x[a] in ['-','—',"=",'_','--']:
                x[a] = 0
            try:
                x[a] = float(x[a])
            except:
                x[a] = 'CHECK'
        
    return x

def check_sum_match(df, tabbed_list):
    
    # Get the integer locations of the tabbed_list items
    tabbed_indices = df.index.get_indexer_for(tabbed_list)
    tabbed_indices = tabbed_indices[tabbed_indices != -1]  # Remove any -1 (not found) values
    print(tabbed_indices)   
    tabbed_indices = sorted(set(tabbed_indices))

    groups = []
    current_group = [tabbed_indices[0]]
    
    for i in range(1, len(tabbed_indices)):
        if tabbed_indices[i] == tabbed_indices[i-1] + 1:
            current_group.append(tabbed_indices[i])
        else:
            groups.append(current_group)
            current_group = [tabbed_indices[i]]
    
    groups.append(current_group)
    print(groups)

    matched_groups = []
    for group in groups:
    
        start_idx = group[0]
        end_idx = group[-1]
        
        sum_of_consecutive_1 = df.iloc[start_idx:end_idx+1, 0].sum()
        sum_of_consecutive_2 = df.iloc[start_idx:end_idx+1, 1].sum()
        value_before_1 = df.iloc[start_idx-1, 0]
        value_before_2 = df.iloc[start_idx-1, 1]

        if value_before_1 != '' and value_before_2 != '':

        #print(sum_of_consecutive_1,sum_of_consecutive_2,value_before_1,value_before_2)

            if np.isclose(sum_of_consecutive_1, value_before_1) and \
            np.isclose(sum_of_consecutive_2, value_before_2) :
                matched_groups.append(group)
    
    return matched_groups

def join_by_commas(x):
    i = 0
    l = len(x)
    while i < l:

        if x[i][0] == '.' \
        or x[i][0] == ',' :
            if i < len(x):
                x = x[:i-1] + [str(x[i-1]+x[i])] + x[i+1:]
                i += 0
                l = len(x)
            else:
                x = x[:i-1] + [str(x[i-1]+x[i])]
                break
        i += 1
    return x



def join_brackets(x):
    
    brackets_list = [] #Joining up lines where bracketed terms have been split
    new_line = []
    for a in range(len(x)):    
        if '(' in x[a]:
            if ')' in x[a]:
                pass 
            elif any(')' in term for term in x[a:]):    #Finding brackets that haven't been closed
                    next_closed_br = [i for i, x in enumerate(')' in term for term in x[a:]) if x][0] 
                    if any('(' in term for term in x[a+1:a+next_closed_br+1]) == False:
                        brackets_list.append([a, a + next_closed_br+1])
    start = 0
    if brackets_list != []:
        for i in range(len(brackets_list)):
            new_line += x[start:brackets_list[i][0]]
            new_line += [''.join(x[brackets_list[i][0]:brackets_list[i][1]])]

            start = brackets_list[i][1]
        new_line += x[start:] 
    else:
        new_line = x
   
    return new_line

def PDF_to_df(doc,x,language='eng'):
 
    doc = pymupdf.open(doc) # open a document
    extracted = unicodedata.normalize("NFKD", doc[x].get_text().replace(u'\u200b','')).splitlines()
    
     # Remove empty lines, but check if this step is over-filtering
    extracted = [line for line in extracted if line.strip() != '']
    
    #while ' ' in extracted:
        #extracted.remove(' ')
    
    
    extracted = pd.Series(extracted)
    for i in range(len(extracted)):
        extracted[i] = extracted[i].strip()
    
    extracted = list(extracted)
    
    array = []
    list_2 = []
    
    for a in extracted:
        if a == "-":
            a = '0' 
        if a == "—":
            a= '0'
        if a == "=":
            a= '0'
        if a == "_":
            a= '0'
        if a == "--":
            a= '0'
        else:
 
 
            if any(char.isdigit() for char in a) == False:
            
                if len(a) >= 3: #not sure, it picks up lots of artifacts
                    array.append(list_2)
                    list_2 = []
                    list_2.append(a)
            else:
                list_2.append(a)

    array.append(list_2) 
    array = array[1:]
    array_1 = np.array(array, dtype=object)

    if max([len(x) for x in array]) <= 2: # this is a band-aid, I really don't understand the problem here, but it shouldn't affect any actual statements/notes as they'll have two years of data  
        return pd.DataFrame(' '.join(x) for x in array_1), pd.DataFrame(' '.join(x) for x in array_1)  # I think the problem is something to do with the first/last lines not being picked up/closed    
    
    df = pd.DataFrame(array_1)
    index = []
    drop_list = []
    for line in range(len(array)):
        a = array[line][0]
        if line > 1 and line <len(array)-1 and a != '':
            if a[0].islower() and len(array[line-1])==1:                
                if array[line-1][0][0].islower() and len(array[line-1]) ==1:    
                    index.append(str(array[line-2][0] +' '+array[line-1][0] +' '+ a))
                    drop_list.append(array[line-2][0])
                    drop_list.append(array[line-1][0])              
                elif array[line+1][0][0].islower() and len(array[line])==1: #two consecutive empty lines
                    index.append(a)        
                else:    
                    index.append(str(array[line-1][0] +' '+ a))
                    drop_list.append(array[line-1][0])
            else:
                index.append(a)
        else:
            index.append(a)

    df.index = index 

    for drop_line in drop_list:
        try: 
            df = df.drop(drop_line)
        except:
            pass

    new_index = []
    for item in df.index:
        new_index.append(item.strip())    
    df.index = new_index
    notes = []
    
    list_notes_s = [str(i) for i in range(2,36)[::1]]
    list_notes_neg = [int(i) for i in range(-36,0)] 

    for i in range(len(df[0])):
        if len(df[0][i]) > 3:
            a = df[0][i][1]  # is first item a note? 
            if a in list_notes_s:
                df[0][i].remove(a)
                notes.append(a) 
            elif process_num([a])[0] in list_notes_neg:
                df[0][i].remove(a)   
            elif process_num([a]) == ['CHECK']:
                    df[0][i].remove(a)
                    notes.append(a)
            elif ',' not in a:
                try:
                    if float(a) > 0 and float(a) < 40 and len(a) <=4: #This has issues with financials containing decimals
                        df[0][i].remove(a)
                        notes.append(a)            
                    else:
                        notes.append('')
                except:
                    notes.append('')
            elif ',' in a and (len(a) <4 or len(a.split(',')[-1])<3) and process_num([a]) != [0]: #wont work with , decimal points
                df[0][i].remove(a)
                notes.append(a)

            else:
                notes.append('')
        else:
            notes.append('')

    year_1 = []
    for a in range(len(df[0])):
        try:
            year_1.append(df[0][a][1])
        except:
            year_1.append('')
    year_2 = []
    for a in range(len(df[0])):
        try:
            year_2.append(df[0][a][2])
        except:
            year_2.append("")
 #sometimes other stuff gets picked up and year 2 gets pushed into the
 
    overflow_1 = []
    for a in range(len(df[0])):
        try:
            overflow_1.append(df[0][a][3])
        except:
            overflow_1.append("")
    overflow_2 = []
    for a in range(len(df[0])):
        try:
            overflow_2.append(df[0][a][4])
        except:
            overflow_2.append("")
    
    
    year_1_raw = year_1.copy()
    year_1 = process_num(year_1)
    
    year_2_raw = year_2.copy()
    year_2 = process_num(year_2)
    
    overflow_1_raw = overflow_1.copy()
    overflow_1 = process_num(overflow_1)
    
    overflow_2_raw = overflow_2.copy()
    overflow_2 = process_num(overflow_2)
 
    df['year_1'] = year_1
    df['year_2'] = year_2
    df['Overflow_1'] = overflow_1
    df['Overflow_2'] = overflow_2
    df['notes'] = notes
    df['year_1_original'] = year_1_raw
    df['year_2_original'] = year_2_raw
    df['Overflow_1_original'] = overflow_1_raw
    df['Overflow_2_original'] = overflow_2_raw
    df_unprocessed = pd.DataFrame(pd.DataFrame(' '.join(x) for x in array_1))
    df = df.drop([0], axis=1)
        
    #display(df)
    return df,   df_unprocessed 

def Image_to_df(image,language='eng'):

    if language == 'eng':
        PLString = pytesseract.image_to_string(image, lang=language, config="--psm 6 -c \
        tessedit_char_whitelist='&)(0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMONPQRSTUVWXYZ,-_ —=./'")
    else:
         PLString = pytesseract.image_to_string(image, lang=language, config="--psm 6")
    PL = PLString.splitlines()
    while '' in PL:
        PL.remove('')   
    table = []

    for line in PL:
        heading = ''
        heading_count = 0
        for word in line.split(' '):

            if word in ['-','—',"=",'_','--']:
                if heading_count ==0:   #This is a sloppy workaround because they sometimes start a line with -
                    heading += word     #Not sure if it always works
                    heading_count+=1
                                    
            elif any(char.isdigit() for char in word) == False: 
                if len(word) > 1:    #This is new to remove artefacts, would cause an issue if it read a number as a single non-numeric 
                    if heading != '': # would also miss a single letter word
                        heading+= ' '
                    heading += word
                heading_count+=1
        new_line = [heading]
        for num in line.split(' ')[heading_count:]:
            
            if any(char.isdigit() for char in num) == True:
                new_line.append(num) #want to stop picking up garbage from photocopies
                
            if num in ['-','—',"=",'_','--']:
                new_line.append(num) #need these non numbers

        table.append(new_line)

    table = table[1:] #don't remember what this is for 
    table_1 = np.array(table, dtype=object)
    
    table_2 = [] #New There was an issue with this bit so i removed it.
    for line in table_1: #
        try:
            table_2.append([line[0]] + join_by_commas(join_brackets(line[1:]))) #running both new functions
        except:
            table_2.append(line)
    table_2 = np.array(table_2, dtype=object) #
    
    df = pd.DataFrame(table_2) #back to table 2 for some reason

    index = []
    drop_list =[]

    # Combining line items when the first is empty and the second one begins in lowercase #NEW

    for line in range(len(table)):
            a = table[line][0]

            if line > 0 and a != '':
                if a[0].islower() and len(table[line-1])==1 and table[line] != a:
                    index.append(str(table[line-1][0] +' '+ a))
                    drop_list.append(table[line-1][0])

                else:
                    index.append(a)
            else:
                index.append(a)
  
    df.index = index
    for drop_line in drop_list:
        try: 
            df = df.drop(drop_line)
        except:
            pass
    notes = []
    
    list_notes_s = [str(i) for i in range(1,36)[::1]]
    list_notes_neg = [int(i) for i in range(-36,0)]

    for i in range(len(df[0])):
        if len(df[0][i]) > 3:  
            a = df[0][i][1]  # is first item a note? 
            if a in list_notes_s:
                df[0][i].remove(a)
                notes.append(a)
            elif process_num([a])[0] in list_notes_neg:
                df[0][i].remove(a)
                notes.append(a)
           
            elif process_num([a]) == ['CHECK']:
                    df[0][i].remove(a)
                    notes.append(a)
            elif ',' not in a:
                try:
                    if float(a) < 40 and len(a) <=4: #This has issues with financials containing decimals
                        df[0][i].remove(a)
                        notes.append(a)                               
                    else:
                        notes.append('')
                except:
                    notes.append('')
            elif ',' in a and (len(a) <4 or len(a.split(',')[-1])<3): #wont work with , decimal points
                df[0][i].remove(a)
                notes.append(a)
            else:
                notes.append('')
        else:
            notes.append('')

    year_1 = []
    for a in range(len(df[0])):
        try:
            year_1.append(df[0][a][1])
        except:
            year_1.append('')
    year_2 = []
    for a in range(len(df[0])):
        try:
            year_2.append(df[0][a][2])
        except:
            year_2.append("")
    
    
    
    #sometimes other stuff gets picked up and year 2 gets pushed into the 3rd column

    overflow_1 = []
    for a in range(len(df[0])):
        try:
            overflow_1.append(df[0][a][3])
        except:
            overflow_1.append("")
    overflow_2 = []
    for a in range(len(df[0])):
        try:
            overflow_2.append(df[0][a][4])
        except:
            overflow_2.append("")
            
            
    year_1_raw = year_1.copy()

    year_1 = process_num(year_1)
                
    year_2_raw = year_2.copy()
    
    year_2 = process_num(year_2)
    
    overflow_1_raw = overflow_1.copy()
    
    overflow_1 = process_num(overflow_1)
    
    overflow_2_raw = overflow_2.copy()
    
    overflow_2 = process_num(overflow_2)
    
    df['year_1'] = year_1
    df['year_2'] = year_2
    df['Overflow_1'] = overflow_1
    df['Overflow_2'] = overflow_2
    df['notes'] = notes
    df['year_1_original'] = year_1_raw
    df['year_2_original'] = year_2_raw
    df['Overflow_1_original'] = overflow_1_raw
    df['Overflow_2_original'] = overflow_2_raw

    df_unprocessed = pd.DataFrame(' '.join(x) for x in table)
    
    df = df.drop([0], axis=1)

    return df, df_unprocessed

def convert_pdf(PDF_name, Output_file, language, pages=None):
    doc = pymupdf.open(PDF_name)
    workbook = Workbook()
    
    # Define styles

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    def apply_format1(cell):
        cell.number_format = '#,##0_ ;-#,##0'
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(size=14)
    
    def apply_format2(cell):
        cell.alignment = Alignment(horizontal='left')
        cell.font = Font(size=16)

    def apply_index_format(cell):
        cell.font = Font(size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
    # If 'pages' is None, process all pages; otherwise, process the selected pages
    if pages is None:
        pages_to_process = range(len(doc))  # Process all pages
    else:
        pages_to_process = pages  # Use the user-provided list/range of pages
        
    for page_no in pages_to_process:
        page = doc[page_no]
        extracted = page.get_text().splitlines()
        if extracted:
            try:
                result = PDF_to_df(doc, page_no, language=language)
                df = result[0]
            except:
                df = pd.DataFrame(extracted)
                result = []
        else:
            pix = page.get_pixmap(dpi=300)
            bytes = np.frombuffer(pix.samples, dtype=np.uint8)
            img = bytes.reshape(pix.height, pix.width, pix.n)
            result = Image_to_df(img, language=language)
            df = result[0]

        # Create worksheets
        ws1 = workbook.create_sheet(title=f'Page{page_no+1}')
        ws2 = workbook.create_sheet(title=f'Page_r{page_no+1}')

        
        # Write data to worksheets
        for r_idx, row in enumerate(dataframe_to_rows(df.reset_index(), index=False, header=True), 1):
            for c_idx, value in enumerate(row, 2):  # Start from column 2
                cell = ws1.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Make row 1 bold
                    cell.font = Font(size = 12,bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add index to ws1
        for i in range(1, ws1.max_row + 1):
            cell = ws1.cell(row=i, column=1, value=i-1 if i > 1 else '')
            apply_index_format(cell)
        if result !=[]:
        # Write data to ws2 and add index
            for r_idx, row in enumerate(dataframe_to_rows(result[1], index=False, header=False), 1):
                # Add index
                index_cell = ws2.cell(row=r_idx, column=1, value=r_idx-1)
                apply_index_format(index_cell)
                # Add data
                for c_idx, value in enumerate(row, 2):
                    try:
                        ws2.cell(row=r_idx, column=c_idx, value=value)
                    except:
                        ws2.cell(row=r_idx, column=c_idx, value= 'character error')
        
        # Insert new column only if language is not 'eng'
        
        columns = ['A','B','C','D','E','F','G','H','I','J','K','L','M']
        if language != 'eng':
            ws1.insert_cols(2)
            col_adj = 1
            cell = ws1.cell(row = 1, column = 2, value = 
                     f"vvvv To translate from {language}, select the relevant rows in column c, right click and select translate, on the translation tab select {language} and copy the output. \
To paste, left click and select special paste -> paste as Unicode. This will maintain the lines")
            cell.font = Font(size = 12,bold=True)
            ws1.column_dimensions['B'].width = 60
        else:
            col_adj = 0

        # Apply styles
        for col in range(2,6):  # Shifted one column to the right
            for cell in ws1[columns[col+col_adj]][1:]:  # Skip the first row (header)
                apply_format1(cell)
            ws1.column_dimensions[columns[col+col_adj]].width = 20
        
        for cell in ws1[columns[1+col_adj]][1:]:
            apply_format2(cell)
        
        for cell in ws1[columns[6+col_adj]][1:]:  # Shifted one column to the right
            apply_format2(cell)
        
        for cell in ws2[columns[1+col_adj]]:  # Shifted one column to the right
            apply_format2(cell)

        # Set column widths
        ws1.column_dimensions[columns[0]].width = 5  # Width for index column
        ws1.column_dimensions[columns[1+col_adj]].width = 60
        ws2.column_dimensions['A'].width = 5
        ws2.column_dimensions['B'].width = 5

        # Set row height
        for row in ws1.iter_rows():
            ws1.row_dimensions[row[0].row].height = 20

    # Remove default sheet 
    if 'Sheet' in workbook.sheetnames and len(workbook.sheetnames) > 1:
        workbook.remove(workbook['Sheet'])

    # Save the workbook
        workbook.save(Output_file)



st.title('DAS F-A-S-T')
st.header('PDF to Excel Convertor')

# Toggle button for showing/hiding the introduction text
if "show_intro" not in st.session_state:
    st.session_state.show_intro = True

if st.button("Show/Hide Instructions"):
    st.session_state.show_intro = not st.session_state.show_intro

if st.session_state.show_intro:
    st.text("""This tool aids the spreading process by converting PDFs into an Excel format. This is useful when the PDF needs translating, or the data in the PDF can't be copied and pasted.
    
    Please choose a PDF file by dragging and dropping it into the drag and drop area. Depending on the size and quality of the file, the processing time might vary from 1 to 5 minutes.
    
    Disclaimer - This product is still currently in Beta and may experience bugs. Thank you for your patience when testing this new feature!""")

# Toggle button for the link to bug report
st.link_button("Report a BUG", "https://ebrd0-my.sharepoint.com/:o:/g/personal/xuc_ebrd_com/Esrs05rTgZZGtageYtBx6JwB3fw8zeMg20fq3bogJk0EjA",icon="🚨")

# File uploader
uploaded_file = st.file_uploader("Choose a PDF file", type="pdf")

if uploaded_file:
    language = st.selectbox("Choose the language of the file", ('eng', 'rus', 'tur','fra'))
    # Page Range Input
    page_input = st.text_input("Enter page numbers (e.g., '1,2,3' or '1-3')", value="")
    if page_input:
        try:
            pages = []
            page_ranges = page_input.split(',')
            for range_str in page_ranges:
                if '-' in range_str:
                    start, end = map(int, range_str.split('-'))
                    pages.extend(range(start-1, end))
                else:
                    pages.append(int(range_str)-1)
        except ValueError:
            st.error("Invalid page range format. Please use commas or hyphens to separate page numbers.")
    else:
        pages = None  # If no pages specified, process all pages

    if st.button("Process PDF"):
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            tmp_file.write(uploaded_file.read())
            pdf_path = tmp_file.name
            
        st.text("Processing PDF...")
    
        # Set the output filename
        output_filename = f"{uploaded_file.name.replace('.pdf', '')}.xlsx"
    
        # Process the PDF
        try:
            convert_pdf(pdf_path, output_filename, language=language, pages=pages)
            st.success("PDF processed successfully!")
    
            with open(output_filename, 'rb') as f:
                st.download_button("Download Excel", f, file_name=output_filename)
        except Exception as e:
            st.error(f"An error occurred while processing the PDF: {str(e)}")
